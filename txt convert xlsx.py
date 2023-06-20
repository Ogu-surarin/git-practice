import os
import shutil
import time
from pathlib import Path
import openpyxl
import csv

wb = openpyxl.load_workbook("oroginal.xlsx")
ws = wb["original"]

move_from = "reading_data"
move_path = "done_reading"
submission_path = ""

for file in Path(move_from).glob("*.txt"):
    read_file = open(file, encoding="utf-8")
    read_row = csv.reader(read_file)
    print("data reading clear.")

    new_ws = wb.copy_worksheet(ws)
    new_ws.title = "No." + str(len(wb.worksheets) - 1 )
    start_r = 21
    for row_data in read_row:
        for data in row_data:
            new_ws.cell(row=start_r, column=2).value = data
            start_r += 1
        print("data copy clear.")
    read_file.close()
    shutil.move(file, move_path)
    print("file moved clear.")

wb.save("sample.xlsx")
wb.close()

subm_wb = openpyxl.load_workbook("original.xlsx")
subm_wb.save("submission.xlsx")
subm_wb.close()

time.sleep(1)
print("")
print("All process Complete.")
time.sleep(0.5)